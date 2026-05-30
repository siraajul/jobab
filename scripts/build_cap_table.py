"""
Generate a clean pre-seed cap table template as XLSX.

Run:
    python3 scripts/build_cap_table.py

Output:
    docs/business/dist/jobab-cap-table.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "docs" / "business" / "dist" / "jobab-cap-table.xlsx"

GREEN_HEX = "1F6E47"
GREEN_DARK_HEX = "154D31"
CREAM_HEX = "FAF7F2"
CREAM_2_HEX = "F4EFE6"
INK_HEX = "2A2722"
INK_2_HEX = "6C645A"
AMBER_HEX = "9A6B0E"
WHITE_HEX = "FFFFFF"


def F(c): return PatternFill("solid", fgColor=c)
def Bs(c="E0E0E0"): return Side(border_style="thin", color=c)


BORDER = Border(left=Bs(), right=Bs(), top=Bs(), bottom=Bs())
HDR_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
HDR_FILL = F(GREEN_HEX)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=GREEN_DARK_HEX)
SECTION_FILL = F(CREAM_2_HEX)
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=GREEN_DARK_HEX)
SUB_FONT = Font(name="Calibri", size=10, italic=True, color=INK_2_HEX)
NUM_FONT = Font(name="Calibri", size=10, color=INK_HEX)


wb = Workbook()

# ─── Sheet 1: Current cap table (pre-seed, pre-investment) ──────────────
ws = wb.active
ws.title = "Current"

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 28

ws["A1"] = "Jobab — cap table"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:E1")
ws["A2"] = "Pre-seed snapshot. Update when anything changes — every investor will ask for the latest version."
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:E2")
ws["A3"] = "Last updated: [insert date]"
ws["A3"].font = SUB_FONT

# Section: Founders
ws["A5"] = "FOUNDERS"
ws["A5"].font = SECTION_FONT
ws["A5"].fill = SECTION_FILL
ws.merge_cells("A5:E5")

headers = ["Holder", "Role", "Shares", "% (fully diluted)", "Vesting"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=6, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

# Default: solo founder owns 80% on a 10M share base, 20% reserved for ESOP
ws.append(["[Founder name]", "Founder, CEO", 8000000, "=C7/$C$25", "4-year vest, 1-year cliff"])

# Section: Co-founders (if any)
ws["A9"] = "CO-FOUNDERS (if any)"
ws["A9"].font = SECTION_FONT
ws["A9"].fill = SECTION_FILL
ws.merge_cells("A9:E9")
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=10, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
ws.append(["[Co-founder name, if any]", "[Role]", 0, "=C11/$C$25", ""])

# Section: ESOP
ws["A13"] = "EMPLOYEE STOCK OPTION POOL (ESOP)"
ws["A13"].font = SECTION_FONT
ws["A13"].fill = SECTION_FILL
ws.merge_cells("A13:E13")
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=14, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
ws.append(["ESOP pool (unallocated)", "Reserved for hires", 2000000, "=C15/$C$25", "Allocate per hire"])
ws.append(["[Senior engineer]", "Allocated (when hired)", 0, "=C16/$C$25", "4yr / 1yr cliff"])
ws.append(["[Merchant ops]", "Allocated (when hired)", 0, "=C17/$C$25", "4yr / 1yr cliff"])

# Section: Advisors (if any)
ws["A19"] = "ADVISORS (if any)"
ws["A19"].font = SECTION_FONT
ws["A19"].fill = SECTION_FILL
ws.merge_cells("A19:E19")
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=20, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
ws.append(["[Advisor name]", "Strategic advisor", 0, "=C21/$C$25", "2yr vest, 6mo cliff"])

# Section: Investors (post-money — empty until close)
ws["A23"] = "INVESTORS (after first close)"
ws["A23"].font = SECTION_FONT
ws["A23"].fill = SECTION_FILL
ws.merge_cells("A23:E23")
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=24, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

# Total row
ws.cell(row=25, column=1, value="TOTAL (fully diluted)").font = Font(bold=True, color=WHITE_HEX)
ws.cell(row=25, column=1).fill = F(GREEN_HEX)
ws.cell(row=25, column=3, value="=C7+C11+C15+C16+C17+C21").font = Font(bold=True, color=WHITE_HEX)
ws.cell(row=25, column=3).fill = F(GREEN_HEX)
ws.cell(row=25, column=4, value="=SUM(D7,D11,D15,D16,D17,D21)").font = Font(bold=True, color=WHITE_HEX)
ws.cell(row=25, column=4).fill = F(GREEN_HEX)
ws.cell(row=25, column=4).number_format = "0.00%"

# Format share columns + % columns
for row in [7, 11, 15, 16, 17, 21, 25]:
    if ws.cell(row=row, column=3).value is not None and row != 25:
        ws.cell(row=row, column=3).number_format = "#,##0"
        ws.cell(row=row, column=4).number_format = "0.00%"
    ws.cell(row=row, column=3).border = BORDER
    ws.cell(row=row, column=4).border = BORDER


# ─── Sheet 2: Post-seed scenario ─────────────────────────────────────────
post = wb.create_sheet("Post-seed scenario")
post.column_dimensions["A"].width = 30
post.column_dimensions["B"].width = 18
post.column_dimensions["C"].width = 16
post.column_dimensions["D"].width = 16
post.column_dimensions["E"].width = 18

post["A1"] = "Cap table — after $400K seed close"
post["A1"].font = TITLE_FONT
post.merge_cells("A1:E1")
post["A2"] = "Assumes $400K at $4M post-money (10% dilution)."
post["A2"].font = SUB_FONT
post.merge_cells("A2:E2")

# Inputs
post["A4"] = "DEAL TERMS (edit these)"
post["A4"].font = SECTION_FONT
post["A4"].fill = SECTION_FILL
post.merge_cells("A4:B4")

post["A5"] = "Round size (USD)"
post["B5"] = 400000
post["B5"].number_format = '"$"#,##0'
post["A6"] = "Post-money valuation (USD)"
post["B6"] = 4000000
post["B6"].number_format = '"$"#,##0'
post["A7"] = "Pre-money valuation"
post["B7"] = "=B6-B5"
post["B7"].number_format = '"$"#,##0'
post["A8"] = "Investor % (post-money)"
post["B8"] = "=B5/B6"
post["B8"].number_format = "0.00%"
post["A9"] = "Pre-round share count"
post["B9"] = 10000000
post["B9"].number_format = "#,##0"
post["A10"] = "New investor shares"
post["B10"] = "=B9*B8/(1-B8)"
post["B10"].number_format = "#,##0"
post["A11"] = "Total shares post-close"
post["B11"] = "=B9+B10"
post["B11"].number_format = "#,##0"
post["A12"] = "Share price (USD)"
post["B12"] = "=B6/B11"
post["B12"].number_format = '"$"#,##0.0000'

# Outputs
post["A14"] = "POST-CLOSE HOLDINGS"
post["A14"].font = SECTION_FONT
post["A14"].fill = SECTION_FILL
post.merge_cells("A14:E14")

headers = ["Holder", "Shares", "% (post-close)", "$ value", "Notes"]
for j, h in enumerate(headers, start=1):
    c = post.cell(row=15, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

rows = [
    ("Founder", 8000000, "Diluted from 80% → 72%"),
    ("Co-founder (if any)", 0, ""),
    ("ESOP pool", 2000000, "May need top-up to keep ESOP at 20% post-round"),
    ("Lead investor", "=B10", "Per deal terms above"),
]
for i, (name, sh, note) in enumerate(rows, start=16):
    post.cell(row=i, column=1, value=name).font = NUM_FONT
    post.cell(row=i, column=2, value=sh).number_format = "#,##0"
    post.cell(row=i, column=2).font = Font(bold=True, color=GREEN_DARK_HEX)
    post.cell(row=i, column=3, value=f"=B{i}/$B$11").number_format = "0.00%"
    post.cell(row=i, column=4, value=f"=B{i}*$B$12").number_format = '"$"#,##0'
    post.cell(row=i, column=5, value=note).font = SUB_FONT
    for c in range(1, 6):
        post.cell(row=i, column=c).border = BORDER

post.cell(row=20, column=1, value="TOTAL").font = Font(bold=True, color=WHITE_HEX)
post.cell(row=20, column=1).fill = F(GREEN_HEX)
post.cell(row=20, column=2, value="=SUM(B16:B19)").font = Font(bold=True, color=WHITE_HEX)
post.cell(row=20, column=2).fill = F(GREEN_HEX)
post.cell(row=20, column=2).number_format = "#,##0"
post.cell(row=20, column=3, value="=SUM(C16:C19)").font = Font(bold=True, color=WHITE_HEX)
post.cell(row=20, column=3).fill = F(GREEN_HEX)
post.cell(row=20, column=3).number_format = "0.00%"
post.cell(row=20, column=4, value="=B20*B12").font = Font(bold=True, color=WHITE_HEX)
post.cell(row=20, column=4).fill = F(GREEN_HEX)
post.cell(row=20, column=4).number_format = '"$"#,##0'


# ─── Sheet 3: Notes ──────────────────────────────────────────────────────
notes = wb.create_sheet("Notes")
notes.column_dimensions["A"].width = 80
notes["A1"] = "How to use this cap table"
notes["A1"].font = TITLE_FONT
notes["A3"] = "WHAT THIS IS"
notes["A3"].font = SECTION_FONT
notes["A4"] = "A simple pre-seed cap table template. Update it every time anything changes."

bullets = [
    "",
    "BEFORE THE FIRST INVESTOR MEETING",
    "  Fill in [Founder name], any co-founders, and the date.",
    "  If you have advisors with equity already, add them.",
    "  Make sure the totals add to exactly 100%.",
    "",
    "ON THE 'POST-SEED' TAB",
    "  Edit the deal terms (round size + post-money valuation).",
    "  The math handles dilution + ESOP top-up suggestions for you.",
    "  Save a copy when you have a real term sheet.",
    "",
    "WHEN AN INVESTOR ASKS FOR THE CAP TABLE",
    "  Send a clean PDF export of the 'Current' tab.",
    "  Don't share the editable Excel until you've signed an NDA.",
    "  Update the 'Last updated' date before sending.",
    "",
    "VOCABULARY (so you don't get lost in investor jargon)",
    "  Fully diluted: total shares if every option were exercised",
    "  Pre-money: company value BEFORE the new investment",
    "  Post-money: company value AFTER the new investment (pre + new round)",
    "  Dilution: your % goes down when new shares are issued",
    "  ESOP: pool of shares set aside for future employees",
    "  Vesting: shares earned over time, not granted instantly",
    "  Cliff: minimum service period before any shares vest",
    "  Anti-dilution: contractual protection against down rounds",
    "  Liquidation preference: investor gets paid first in an exit",
    "",
    "RED FLAGS TO PUSH BACK ON",
    "  > 2x liquidation preference — anything above 1x non-participating is aggressive",
    "  ESOP top-up coming entirely from founder shares (should be shared with investors)",
    "  Founder vesting reset on funding — your existing time should count",
    "  > 25% dilution at seed — you'll run out of room by Series A",
    "",
    "GET A LAWYER",
    "  Before signing a term sheet, get a Bangladeshi corporate lawyer to review.",
    "  This template is for planning, not for legal commitments.",
]

for i, line in enumerate(bullets, start=5):
    notes.cell(row=i, column=1, value=line)
    if line and not line.startswith(" ") and line.isupper() and line != line.lower():
        notes.cell(row=i, column=1).font = SECTION_FONT
    else:
        notes.cell(row=i, column=1).font = NUM_FONT


OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"✓ wrote {OUT.relative_to(REPO_ROOT)} ({OUT.stat().st_size // 1024} KB)")
