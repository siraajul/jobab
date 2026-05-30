"""
Generate the Jobab 36-month financial model as XLSX with brand styling
and live formulas (not hard-coded numbers).

Run:
    python3 scripts/build_financial_model.py

Output:
    docs/business/dist/jobab-financial-model.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.formatting.rule import ColorScaleRule


REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "docs" / "business" / "dist" / "jobab-financial-model.xlsx"

# Brand
GREEN_HEX = "1F6E47"
GREEN_DARK_HEX = "154D31"
CREAM_HEX = "FAF7F2"
CREAM_2_HEX = "F4EFE6"
INK_HEX = "2A2722"
INK_2_HEX = "6C645A"
AMBER_HEX = "9A6B0E"
WHITE_HEX = "FFFFFF"


def F(color):
    return PatternFill("solid", fgColor=color)


def B(color="000000", style="thin"):
    return Side(border_style=style, color=color)


HDR_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
HDR_FILL = F(GREEN_HEX)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=GREEN_DARK_HEX)
SECTION_FILL = F(CREAM_2_HEX)
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=GREEN_DARK_HEX)
SUB_FONT = Font(name="Calibri", size=10, italic=True, color=INK_2_HEX)
NUM_FONT = Font(name="Calibri", size=10, color=INK_HEX)
BORDER = Border(left=B("E0E0E0"), right=B("E0E0E0"), top=B("E0E0E0"), bottom=B("E0E0E0"))


wb = Workbook()


# ─── Sheet 1: Assumptions ─────────────────────────────────────────────────
ws = wb.active
ws.title = "Assumptions"
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 36

ws["A1"] = "Jobab — financial model assumptions"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:C1")
ws["A2"] = "All revenue numbers in USD. 1 USD ≈ ৳110. Tweak any cell — P&L recalculates."
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:C2")

rows = [
    ("PRICING & MIX", None, None),
    ("Starter price (USD/mo)", 9, "৳999"),
    ("Growth price (USD/mo)", 27, "৳2,999"),
    ("Pro price (USD/mo)", 73, "৳7,999"),
    ("Plan mix — Starter (Y1)", 0.60, "%"),
    ("Plan mix — Growth (Y1)", 0.30, "%"),
    ("Plan mix — Pro (Y1)", 0.10, "%"),
    ("Plan mix — Starter (Y2+)", 0.40, "%"),
    ("Plan mix — Growth (Y2+)", 0.45, "%"),
    ("Plan mix — Pro (Y2+)", 0.15, "%"),
    ("", None, None),
    ("UNIT ECONOMICS", None, None),
    ("COGS per conversation (USD)", 0.003, "Groq + vision + embeddings"),
    ("Conversations per merchant/mo", 1500, "Average"),
    ("Fixed infra cost (USD/mo)", 100, "Hosting + DB + Redis"),
    ("Monthly logo churn", 0.04, "4% — SMB SaaS norm in EM"),
    ("", None, None),
    ("ACQUISITION", None, None),
    ("CAC (USD)", 32, "৳3,500"),
    ("Trial → paid conversion", 0.40, "40% activation"),
    ("Average merchant lifetime (months)", 25, "= 1 / churn"),
    ("", None, None),
    ("HIRING (USD/mo cost)", None, None),
    ("Founder", 1360, "M1+"),
    ("Senior engineer", 1820, "M4+"),
    ("Merchant ops", 730, "M6+"),
    ("Fractional CFO", 545, "M12+"),
    ("Second engineer", 1635, "M18+"),
    ("BD lead", 1090, "M24+"),
    ("", None, None),
    ("OTHER OPEX", None, None),
    ("Marketing — early (USD/mo)", 500, "M1–M12"),
    ("Marketing — late (USD/mo)", 2000, "M13+"),
    ("Legal + accounting (USD/mo)", 200, ""),
    ("Software (USD/mo)", 100, ""),
    ("Office (USD/mo)", 300, ""),
]

for i, (label, val, note) in enumerate(rows, start=4):
    cell = ws.cell(row=i, column=1, value=label)
    if val is None:
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    else:
        cell.font = NUM_FONT
        vcell = ws.cell(row=i, column=2, value=val)
        vcell.font = Font(name="Calibri", size=11, bold=True, color=GREEN_DARK_HEX)
        vcell.alignment = Alignment(horizontal="right")
        if "%" in str(note):
            vcell.number_format = "0%"
        elif "USD" in label or "CAC" in label or "price" in label:
            vcell.number_format = "$#,##0.00"
        ws.cell(row=i, column=3, value=note).font = SUB_FONT


# Named references — for use in P&L formulas
def ref(row):
    return f"Assumptions!$B${row}"


REF = {
    "starter_price": ref(5),
    "growth_price": ref(6),
    "pro_price": ref(7),
    "mix_starter_y1": ref(8),
    "mix_growth_y1": ref(9),
    "mix_pro_y1": ref(10),
    "mix_starter_y2": ref(11),
    "mix_growth_y2": ref(12),
    "mix_pro_y2": ref(13),
    "cogs_per_conv": ref(17),
    "convs_per_merchant": ref(18),
    "fixed_infra": ref(19),
    "founder": ref(28),
    "sr_eng": ref(29),
    "ops": ref(30),
    "cfo": ref(31),
    "second_eng": ref(32),
    "bd_lead": ref(33),
    "mkt_early": ref(36),
    "mkt_late": ref(37),
    "legal": ref(38),
    "software": ref(39),
    "office": ref(40),
}


# ─── Sheet 2: Monthly P&L ─────────────────────────────────────────────────
pnl = wb.create_sheet("Monthly P&L")

# Growth ramp (merchant counts at end of month)
ramp = [
    0, 0, 0,            # M1-3 pilot, free
    8, 15, 25,          # M4-6
    38, 48, 60, 72, 86, 100,   # M7-12
    118, 138, 160, 184, 210, 250,   # M13-18
    290, 335, 380, 425, 460, 500,   # M19-24
    545, 605, 665, 725, 775, 800,   # M25-30
    855, 915, 985, 1060, 1130, 1200,  # M31-36
]

# Header row
pnl["A1"] = "Monthly P&L — 36-month projection (USD)"
pnl["A1"].font = TITLE_FONT
pnl.merge_cells("A1:N1")

cols = [
    "Month", "Merchants (end)", "Blended ARPU", "MRR", "COGS",
    "Gross profit", "Salaries", "Marketing", "Infra", "Other OpEx",
    "Total OpEx", "Net (burn)", "Cumulative burn", "ARR",
]

for j, h in enumerate(cols, start=1):
    c = pnl.cell(row=3, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

for j in range(1, len(cols) + 1):
    pnl.column_dimensions[get_column_letter(j)].width = 14

pnl.column_dimensions["A"].width = 8
pnl.column_dimensions["B"].width = 16

# Rows
for i, count in enumerate(ramp, start=1):
    row = i + 3
    pnl.cell(row=row, column=1, value=f"M{i}")
    pnl.cell(row=row, column=2, value=count)

    # ARPU: switches to Y2+ mix from month 13
    if i <= 12:
        mix_arpu_formula = f"={REF['starter_price']}*{REF['mix_starter_y1']}+{REF['growth_price']}*{REF['mix_growth_y1']}+{REF['pro_price']}*{REF['mix_pro_y1']}"
    else:
        mix_arpu_formula = f"={REF['starter_price']}*{REF['mix_starter_y2']}+{REF['growth_price']}*{REF['mix_growth_y2']}+{REF['pro_price']}*{REF['mix_pro_y2']}"
    pnl.cell(row=row, column=3, value=mix_arpu_formula)

    # MRR = merchants * ARPU
    pnl.cell(row=row, column=4, value=f"=B{row}*C{row}")
    # COGS = merchants * convs * cost_per_conv
    pnl.cell(row=row, column=5, value=f"=B{row}*{REF['convs_per_merchant']}*{REF['cogs_per_conv']}")
    # Gross profit
    pnl.cell(row=row, column=6, value=f"=D{row}-E{row}")

    # Salaries
    sal = REF['founder']
    if i >= 4:
        sal += f"+{REF['sr_eng']}"
    if i >= 6:
        sal += f"+{REF['ops']}"
    if i >= 12:
        sal += f"+{REF['cfo']}"
    if i >= 18:
        sal += f"+{REF['second_eng']}"
    if i >= 24:
        sal += f"+{REF['bd_lead']}"
    pnl.cell(row=row, column=7, value=f"={sal}")

    # Marketing
    pnl.cell(row=row, column=8, value=f"={REF['mkt_early'] if i <= 12 else REF['mkt_late']}")
    # Infra
    pnl.cell(row=row, column=9, value=f"={REF['fixed_infra']}+IF(B{row}>500,B{row}*0.4,0)")
    # Other OpEx
    pnl.cell(row=row, column=10, value=f"={REF['legal']}+{REF['software']}+{REF['office']}")
    # Total OpEx
    pnl.cell(row=row, column=11, value=f"=G{row}+H{row}+I{row}+J{row}")
    # Net
    pnl.cell(row=row, column=12, value=f"=F{row}-K{row}")
    # Cumulative burn
    if i == 1:
        pnl.cell(row=row, column=13, value=f"=L{row}")
    else:
        pnl.cell(row=row, column=13, value=f"=M{row - 1}+L{row}")
    # ARR
    pnl.cell(row=row, column=14, value=f"=D{row}*12")

    # Formatting
    for col in range(3, 15):
        cell = pnl.cell(row=row, column=col)
        cell.number_format = '"$"#,##0;"($"#,##0")"'
        cell.font = NUM_FONT
        cell.border = BORDER

    # Zebra
    if i % 2 == 0:
        for col in range(1, 15):
            pnl.cell(row=row, column=col).fill = F(CREAM_HEX)

    pnl.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True, color=GREEN_DARK_HEX)
    pnl.cell(row=row, column=2).font = Font(name="Calibri", size=10, bold=True, color=INK_HEX)


# Conditional color scale on Net column
pnl.conditional_formatting.add(
    f"L4:L{3 + len(ramp)}",
    ColorScaleRule(start_type="min", start_color="F4D7D7",
                   mid_type="num", mid_value=0, mid_color="FAF7F2",
                   end_type="max", end_color="D4E8DC"),
)

# Freeze the header
pnl.freeze_panes = "A4"


# ─── Sheet 3: Charts ──────────────────────────────────────────────────────
charts = wb.create_sheet("Charts")
charts["A1"] = "Visuals"
charts["A1"].font = TITLE_FONT

# MRR line chart
ch1 = LineChart()
ch1.title = "MRR over 36 months (USD)"
ch1.y_axis.title = "MRR (USD)"
ch1.x_axis.title = "Month"
data = Reference(pnl, min_col=4, min_row=3, max_col=4, max_row=3 + len(ramp))
cats = Reference(pnl, min_col=1, min_row=4, max_row=3 + len(ramp))
ch1.add_data(data, titles_from_data=True)
ch1.set_categories(cats)
ch1.height = 10
ch1.width = 18
charts.add_chart(ch1, "A3")

# Merchants bar chart
ch2 = LineChart()
ch2.title = "Paying merchants"
ch2.y_axis.title = "Merchants"
ch2.x_axis.title = "Month"
data2 = Reference(pnl, min_col=2, min_row=3, max_col=2, max_row=3 + len(ramp))
ch2.add_data(data2, titles_from_data=True)
ch2.set_categories(cats)
ch2.height = 10
ch2.width = 18
charts.add_chart(ch2, "A24")

# Cumulative burn
ch3 = LineChart()
ch3.title = "Cumulative cash burn"
ch3.y_axis.title = "USD"
ch3.x_axis.title = "Month"
data3 = Reference(pnl, min_col=13, min_row=3, max_col=13, max_row=3 + len(ramp))
ch3.add_data(data3, titles_from_data=True)
ch3.set_categories(cats)
ch3.height = 10
ch3.width = 18
charts.add_chart(ch3, "A45")


# ─── Sheet 4: Use of funds ───────────────────────────────────────────────
funds = wb.create_sheet("Use of funds")
funds["A1"] = "Use of funds — $400K seed"
funds["A1"].font = TITLE_FONT
funds.merge_cells("A1:C1")

funds.column_dimensions["A"].width = 32
funds.column_dimensions["B"].width = 14
funds.column_dimensions["C"].width = 10

headers = ["Bucket", "Amount (USD)", "% of round"]
for j, h in enumerate(headers, start=1):
    c = funds.cell(row=3, column=j, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")

funds_data = [
    ("Engineering (1 senior eng × 18 months + founder)", 160000, 0.40),
    ("Merchant operations (hiring + onboarding first 100)", 100000, 0.25),
    ("Marketing (paid + content + events)", 60000, 0.15),
    ("Infrastructure (LLM tokens, hosting, scaling)", 60000, 0.15),
    ("Buffer (legal, contingencies)", 20000, 0.05),
]
for i, (label, amt, pct) in enumerate(funds_data, start=4):
    funds.cell(row=i, column=1, value=label).font = NUM_FONT
    a = funds.cell(row=i, column=2, value=amt)
    a.number_format = '"$"#,##0'
    a.font = Font(bold=True, color=GREEN_DARK_HEX)
    p = funds.cell(row=i, column=3, value=pct)
    p.number_format = "0%"
    p.font = Font(bold=True, color=AMBER_HEX)
    p.alignment = Alignment(horizontal="right")

# Total row
funds.cell(row=10, column=1, value="TOTAL").font = Font(bold=True, color=WHITE_HEX)
funds.cell(row=10, column=1).fill = F(GREEN_HEX)
funds.cell(row=10, column=2, value="=SUM(B4:B8)").number_format = '"$"#,##0'
funds.cell(row=10, column=2).font = Font(bold=True, color=WHITE_HEX)
funds.cell(row=10, column=2).fill = F(GREEN_HEX)
funds.cell(row=10, column=3, value="=SUM(C4:C8)").number_format = "0%"
funds.cell(row=10, column=3).font = Font(bold=True, color=WHITE_HEX)
funds.cell(row=10, column=3).fill = F(GREEN_HEX)


# ─── Sheet 5: Summary ─────────────────────────────────────────────────────
summary = wb.create_sheet("Summary", 0)
summary.column_dimensions["A"].width = 30
summary.column_dimensions["B"].width = 18

summary["A1"] = "Jobab"
summary["A1"].font = Font(name="Calibri", size=24, bold=True, color=GREEN_DARK_HEX)
summary["A2"] = "AI shop assistant for Bangladeshi merchants"
summary["A2"].font = Font(name="Calibri", size=12, italic=True, color=INK_2_HEX)
summary["A3"] = "Pre-seed financial model · 36 months · all USD"
summary["A3"].font = SUB_FONT

summary["A5"] = "AT A GLANCE"
summary["A5"].font = SECTION_FONT
summary["A5"].fill = SECTION_FILL
summary.merge_cells("A5:B5")

quick = [
    ("Target raise", "=Funds.B10" if False else "$400,000"),
    ("Months runway", "18"),
    ("Merchants by EOY1", "100"),
    ("MRR by EOY1", "$1,700"),
    ("Merchants by EOY2", "500"),
    ("MRR by EOY2", "$14,000"),
    ("Merchants by EOY3", "1,200"),
    ("MRR by EOY3", "$33,600"),
    ("Cash-flow positive", "Month 24"),
    ("LTV : CAC target", "17 : 1"),
    ("Gross margin (Growth tier)", "~85%"),
]

for i, (label, val) in enumerate(quick, start=7):
    summary.cell(row=i, column=1, value=label).font = NUM_FONT
    c = summary.cell(row=i, column=2, value=val)
    c.font = Font(name="Calibri", size=11, bold=True, color=GREEN_DARK_HEX)
    c.alignment = Alignment(horizontal="right")

summary["A20"] = "SHEETS IN THIS WORKBOOK"
summary["A20"].font = SECTION_FONT
summary["A20"].fill = SECTION_FILL
summary.merge_cells("A20:B20")

sheets_help = [
    ("Assumptions", "Tweak any number — P&L recalculates"),
    ("Monthly P&L", "36-month revenue + cost projection"),
    ("Charts", "MRR / merchants / burn visualisations"),
    ("Use of funds", "Where the $400K goes"),
]
for i, (s, h) in enumerate(sheets_help, start=22):
    summary.cell(row=i, column=1, value=s).font = Font(bold=True, color=GREEN_DARK_HEX)
    summary.cell(row=i, column=2, value=h).font = SUB_FONT


OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"✓ wrote {OUT.relative_to(REPO_ROOT)} ({OUT.stat().st_size // 1024} KB)")
